import openpyxl

workbook = openpyxl.load_workbook('/Users/parikshithkulkarni/Downloads/example.xlsx')
# print(workbook.sheetnames)
sheet = workbook['Sheet1']
cell = sheet['A1']
# print(cell.value)
# print(str(sheet['A2'].value))

for i in range(1, 8):
    continue
    # print(sheet.cell(row=i, column=2).value)

workbook.close()

wb = openpyxl.Workbook()
print(wb.sheetnames)
sheet1 = wb['Sheet']
sheet1['A1'] = 'hello'
sheet0 = wb.create_sheet(title = 'MySheet', index = 0)
sheet2 = wb.create_sheet()
print(wb.sheetnames)
wb.save('example.xlsx')
wb.close()
